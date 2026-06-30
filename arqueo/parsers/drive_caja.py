"""Parser unificado Drive Caja: lee SOLO las hojas que nos interesan
(CAJA + TORCAL), limita filas a 2000 y comparte el workbook entre los dos
sub-parsers. Mucho más rápido que leer el archivo entero."""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from . import drive_pb, drive_otros


def detect_seccion(filename: str) -> str | None:
    return drive_pb.detect_seccion(filename) or drive_otros.detect_seccion(filename)


def _read_sheet(ws, max_rows: int = 2000) -> list[list]:
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_rows)):
        rows.append(list(row))
        if i > max_rows: break
    return rows


def parse(path: str | Path) -> dict:
    """Devuelve {'pb': df, 'otros': df}. Solo abre el .xlsx una vez y solo
    materializa las 2 hojas relevantes."""
    out = {"pb": pd.DataFrame(), "otros": pd.DataFrame()}
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:
        return out
    try:
        sheets_dict = {}
        for sn in wb.sheetnames:
            su = sn.upper()
            if "TORCAL" in su or ("CAJA" in su and "CONFIG" not in su):
                try:
                    sheets_dict[sn] = _read_sheet(wb[sn])
                except Exception:
                    sheets_dict[sn] = []
    finally:
        wb.close()

    try:
        out["pb"] = drive_pb.parse_sheets(sheets_dict)
    except Exception:
        pass
    try:
        out["otros"] = drive_otros.parse_sheets(sheets_dict)
    except Exception:
        pass
    return out
