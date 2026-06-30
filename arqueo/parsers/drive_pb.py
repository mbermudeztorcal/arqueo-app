"""Parser Drive Permiso B. Usa openpyxl en modo read_only (muy rápido,
tolera fechas mal escritas tipo 'año 20025')."""
from __future__ import annotations
import datetime as dt
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

_VALID_SECS = {"05","07","10","12","15","17","25","28","31","42","47"}


def detect_seccion(filename: str) -> str | None:
    nombre = filename.upper()
    for m in re.finditer(r"\b(\d{2})\b", nombre):
        codigo = m.group(1)
        if codigo in _VALID_SECS:
            return f"S{codigo}"
    return None


def _load_sheets(path: str | Path) -> dict[str, list[list]]:
    """Lee TODAS las hojas como listas de filas. read_only + data_only = 5-10x
    más rápido que pd.read_excel(engine='openpyxl') normal y tolera fechas
    inválidas devolviendo el valor crudo."""
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        return {sn: [list(row) for row in wb[sn].iter_rows(values_only=True)] for sn in wb.sheetnames}
    finally:
        wb.close()


def parse(path: str | Path, sheets: dict | None = None) -> pd.DataFrame:
    if sheets is None:
        sheets = _load_sheets(path)
    return parse_sheets(sheets)


def parse_sheets(sheets: dict) -> pd.DataFrame:
    """Acepta tanto dict de DataFrames (compat) como dict de list-of-lists (rápido)."""
    caja_sheet = None
    for sn in sheets:
        if "CAJA" in sn.upper() and "CONFIG" not in sn.upper():
            caja_sheet = sn
            break
    if not caja_sheet:
        return pd.DataFrame()

    raw = sheets[caja_sheet]
    if isinstance(raw, pd.DataFrame):
        raw = raw.values.tolist()
    if not raw or len(raw) < 2:
        return pd.DataFrame()

    headers = [str(c).strip() if c is not None else "" for c in raw[0]]
    rows_raw = raw[1:]

    def col(*needles):
        for i, h in enumerate(headers):
            hl = h.lower()
            if all(n in hl for n in needles):
                return i
        return None

    i_fecha    = col("fecha")
    i_ingreso  = col("ingreso", "efect")
    i_pago     = col("pago", "efect")
    i_saldo    = col("saldo", "efect")
    i_concepto = col("concepto") if not col("concepto","factur") else col("concepto")
    i_cliente  = col("cliente")
    i_notas    = col("nota")

    if i_fecha is None or i_ingreso is None:
        return pd.DataFrame()

    rows = []
    for r in rows_raw:
        if r is None: continue
        f = r[i_fecha] if i_fecha < len(r) else None
        if not isinstance(f, (dt.datetime, dt.date)):
            continue
        if isinstance(f, dt.datetime): fdate = f.date()
        else: fdate = f
        def _f(idx):
            if idx is None or idx >= len(r): return 0.0
            v = r[idx]
            try: return float(v) if v is not None else 0.0
            except (TypeError, ValueError): return 0.0
        def _v(idx):
            if idx is None or idx >= len(r): return None
            return r[idx]
        rows.append({
            "fecha": fdate,
            "ingreso": _f(i_ingreso), "pago": _f(i_pago),
            "saldo": _f(i_saldo) if i_saldo is not None else None,
            "concepto": _v(i_concepto), "cliente": _v(i_cliente), "notas": _v(i_notas),
        })
    return pd.DataFrame(rows)


def saldos_por_dia(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "saldo" not in df.columns:
        return pd.DataFrame(columns=["fecha","saldo_cierre"])
    df_clean = df[df["saldo"].notna()]
    if df_clean.empty:
        return pd.DataFrame(columns=["fecha","saldo_cierre"])
    return df_clean.groupby("fecha").agg(saldo_cierre=("saldo","last")).reset_index()
