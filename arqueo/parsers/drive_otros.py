"""Parser Drive Otros Permisos (hoja CAJA TORCAL) con openpyxl read_only."""
from __future__ import annotations
import datetime as dt
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


def detect_seccion(filename: str) -> str | None:
    m = re.search(r"S(\d{2})", filename)
    if m: return f"S{m.group(1)}"
    m2 = re.search(r"SEC\.?\s*(\d{2})", filename, re.IGNORECASE)
    if m2: return f"S{m2.group(1)}"
    return None


def _load_sheets(path: str | Path) -> dict[str, list[list]]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        return {sn: [list(row) for row in wb[sn].iter_rows(values_only=True)] for sn in wb.sheetnames}
    finally:
        wb.close()


def parse(path: str | Path, sheets: dict | None = None) -> pd.DataFrame:
    if sheets is None:
        sheets = _load_sheets(path)
    return parse_sheets(sheets)


def parse_sheets(sheets: dict) -> pd.DataFrame:
    torcal_sheet = next((s for s in sheets if "TORCAL" in s.upper()), None)
    if not torcal_sheet:
        return pd.DataFrame()
    raw = sheets[torcal_sheet]
    if isinstance(raw, pd.DataFrame):
        raw = raw.values.tolist()
    if not raw or len(raw) < 2:
        return pd.DataFrame()

    headers = [str(c).strip() if c is not None else "" for c in raw[0]]
    use_positional = not any("fecha" in h.lower() for h in headers)

    def col(*needles):
        for i, h in enumerate(headers):
            hl = h.lower()
            if all(n in hl for n in needles):
                return i
        return None

    rows = []
    for r in raw[1:]:
        if r is None: continue
        if use_positional:
            f = r[0] if len(r) > 0 else None
            c_cobro = r[3] if len(r) > 3 else None
            saldo = r[4] if len(r) > 4 else None
            obs = r[5] if len(r) > 5 else None
        else:
            i_f = col("fecha"); i_c = col("cobro","efect")
            i_s = col("saldo","efect"); i_o = col("observ")
            if i_f is None or i_c is None: continue
            f = r[i_f] if i_f < len(r) else None
            c_cobro = r[i_c] if i_c < len(r) else None
            saldo = r[i_s] if i_s is not None and i_s < len(r) else None
            obs = r[i_o] if i_o is not None and i_o < len(r) else None
        if not isinstance(f, (dt.datetime, dt.date)):
            continue
        if isinstance(f, dt.datetime): fdate = f.date()
        else: fdate = f
        try: c_cobro_f = float(c_cobro) if c_cobro is not None else 0.0
        except (TypeError, ValueError): c_cobro_f = 0.0
        try: saldo_f = float(saldo) if saldo is not None else None
        except (TypeError, ValueError): saldo_f = None
        rows.append({"fecha": fdate, "cobro_efectivo": c_cobro_f,
                     "saldo": saldo_f, "observaciones": obs})
    return pd.DataFrame(rows)


def saldos_por_dia(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "saldo" not in df.columns:
        return pd.DataFrame(columns=["fecha","saldo_cierre"])
    df_clean = df[df["saldo"].notna()]
    if df_clean.empty:
        return pd.DataFrame(columns=["fecha","saldo_cierre"])
    return df_clean.groupby("fecha").agg(saldo_cierre=("saldo","last")).reset_index()
